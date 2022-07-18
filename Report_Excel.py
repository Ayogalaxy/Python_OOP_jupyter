import pandas as pd
import glob, os
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

files = os.listdir()

combined = pd.DataFrame()

for file in files:
    if file.endswith(".xlsx"):
        #        print(file)
        df = pd.read_excel(file)
        combined = pd.concat([combined, df], ignore_index=True)
print(combined)

# combined.to_excel('combined.xlsx', index=False)

summary = pd.pivot_table(
    data=combined,
    index="Salesperson",
    values='Amount',
    aggfunc="sum"
)

summary.to_excel('Summary.xlsx',)

wb = load_workbook('Summary.xlsx')
ws = wb['Sheet1']

ws.insert_rows(0, 3)
ws['A1'].value = 'Sales by Salesperson'
ws['A2'].value = 'datagy.io'

ws['A1'].style = 'Title'
ws['A2'].style = 'Headline 2'

for cell in range(5, ws.max_row+1):
    ws[f"B{cell}"].style = 'Currency'

data = Reference(ws, min_col=2, max_col=2, min_row=5, max_row=ws.max_row)
categories = Reference(ws, min_col=1, max_col=1, min_row=5, max_row=ws.max_row)

chart = BarChart()
chart.add_data(data)
chart.set_categories(categories)
chart.title = 'Sales by Salesperson'
ws.add_chart(chart, anchor='F4')

wb.save('Summary.xlsx')

